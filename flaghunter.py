#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import re
import time
import mmap
import yaml
import base64
import binascii
from datetime import datetime
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import openpyxl
from openpyxl.utils import get_column_letter

CONFIG_FILE = 'config.yml'
LOG_FILE = 'found_flags.log'

DEFAULT_CONFIG = {
    'flag_rules': [{'prefix': 'flag'}, {'prefix': 'ctf'}, {'prefix': 'CTF'}, {'prefix': 'FLAG'}],
    'scan_cooldown': 10,
    'max_file_size_mb': 100,
    'context_bytes': 40,
    'line_search_rules': [
        {'name': '包含 - 和 }', 'includes': ['-', '}']}
    ]
}

class Colors:
    GREEN = '\033[92m'
    CYAN = '\033[96m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'

IS_WIN = sys.platform.startswith('win')

def banner():
    print(f"""{Colors.GREEN}{Colors.BOLD}
==================================================
   ███████╗██╗      █████╗  ██████╗      ██╗  ██╗██╗   ██╗███╗   ██╗████████╗███████╗██████╗ 
   ██╔════╝██║     ██╔══██╗██╔════╝      ██║  ██║██║   ██║████╗  ██║╚══██╔══╝██╔════╝██╔══██╗
   █████╗  ██║     ███████║██║  ███╗     ███████║██║   ██║██╔██╗ ██║   ██║   █████╗  ██████╔╝
   ██╔══╝  ██║     ██╔══██║██║   ██║     ██╔══██║██║   ██║██║╚██╗██║   ██║   ██╔══╝  ██╔══██╗
   ██║     ███████╗██║  ██║╚██████╔╝     ██║  ██║╚██████╔╝██║ ╚████║   ██║   ███████╗██║  ██║
   ╚═╝     ╚══════╝╚═╝  ╚═╝ ╚═════╝      ╚═╝  ╚═╝ ╚═════╝ ╚═╝  ╚═══╝   ╚═╝   ╚══════╝╚═╝  ╚═╝
==================================================
{Colors.CYAN}{Colors.BOLD}                   Coded By {Colors.ENDC}{Colors.CYAN}Hx0战队{Colors.ENDC}
""")

def get_time():
    return datetime.now().strftime("%H:%M:%S")

def log_info(msg):
    color = Colors.CYAN if not IS_WIN else ""
    print(f"{color}[{get_time()}] [INFO] {msg}{Colors.ENDC}")

def log_warn(msg):
    color = Colors.YELLOW if not IS_WIN else ""
    print(f"{color}[{get_time()}] [WARN] {msg}{Colors.ENDC}")

def log_error(msg):
    color = Colors.RED if not IS_WIN else ""
    print(f"{color}[{get_time()}] [ERROR] {msg}{Colors.ENDC}")

def log_flag_to_file(text):
    try:
        with open(LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(text + "\n\n")
    except Exception as e:
        log_warn(f"无法写入日志文件: {e}")

def load_or_create_config():
    path = os.path.abspath(CONFIG_FILE)
    if not os.path.exists(path):
        with open(path, 'w', encoding='utf-8') as f:
            yaml.dump(DEFAULT_CONFIG, f, allow_unicode=True)
        log_info(f"未找到配置文件，已创建默认 {CONFIG_FILE}")
        return DEFAULT_CONFIG
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f)
        if not data:
            raise ValueError("配置为空")
        return {**DEFAULT_CONFIG, **data}
    except Exception as e:
        log_warn(f"加载配置失败，使用默认配置: {e}")
        return DEFAULT_CONFIG

def build_regex_from_rules(rules):
    patterns = []
    for rule in rules:
        prefix = rule.get('prefix', '')
        if not prefix:
            continue
        try:
            # 明文格式 - 更宽松的匹配，允许任何字符直到闭合大括号
            patterns.append((f"Plaintext ({prefix})", re.compile(prefix.encode() + b'\\{[\\s\\S]*?\\}', re.DOTALL)))
            # Base64 格式
            b64 = base64.b64encode((prefix + "{").encode()).rstrip(b'=')
            patterns.append((f"Base64 ({prefix})", re.compile(re.escape(b64) + b'[A-Za-z0-9+/=]{10,}')))
            # Hex 格式
            hex_prefix = binascii.hexlify((prefix + "{").encode())
            patterns.append((f"Hex ({prefix})", re.compile(re.escape(hex_prefix) + b'[0-9a-fA-F]{10,}')))
            # 直接匹配前缀，用于二进制文件
            patterns.append((f"Raw ({prefix})", re.compile(prefix.encode() + b'.{10,}')))
        except Exception as e:
            log_warn(f"规则错误: {e}")
    return patterns

flag_counter = 0
scanned_files = {}
found_flags = []
found_lines = []

def _report_flag(buffer, file_path, context_bytes, rules):
    global flag_counter
    for name, regex in rules:
        for match in regex.finditer(buffer):
            flag_counter += 1
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            found = match.group(0)
            try:
                flag_text = found.decode('utf-8')
            except:
                flag_text = repr(found)

            offset = match.start()
            line_num = 1 + buffer[:offset].count(b'\n')

            ctx_start = max(0, offset - context_bytes)
            ctx_end = min(len(buffer), match.end() + context_bytes)
            ctx = buffer[ctx_start:ctx_end].decode('utf-8', errors='replace')

            decoded = None
            if name.startswith("Base64"):
                try:
                    d = found
                    pad = len(d) % 4
                    if pad:
                        d += b'=' * (4 - pad)
                    decoded = base64.b64decode(d).decode('utf-8', errors='replace')
                except:
                    pass
            elif name.startswith("Hex"):
                try:
                    d = found
                    if len(d) % 2:
                        d = d[:-1]
                    decoded = binascii.unhexlify(d).decode('utf-8', errors='replace')
                except:
                    pass

            border = f"{Colors.GREEN}{'=' * 60}{Colors.ENDC}"
            print(f"\n{border}")
            print(f"{Colors.BOLD}{Colors.GREEN}[ FLAG FOUND #{flag_counter} ]{Colors.ENDC}")
            print(f"{Colors.CYAN}时间:{Colors.ENDC} {timestamp}")
            print(f"{Colors.CYAN}文件:{Colors.ENDC} {file_path}")
            print(f"{Colors.CYAN}类型:{Colors.ENDC} {name}")
            print(f"{Colors.CYAN}行号:{Colors.ENDC} {line_num}")
            print(f"{Colors.CYAN}偏移:{Colors.ENDC} {offset}")
            print(f"{Colors.CYAN}Flag:{Colors.ENDC} {Colors.RED}{flag_text}{Colors.ENDC}")
            if decoded:
                print(f"{Colors.CYAN}解码:{Colors.ENDC} {Colors.RED}{decoded}{Colors.ENDC}")
            print(f"{Colors.CYAN}上下文:{Colors.ENDC}")
            print(f"{Colors.YELLOW}{ctx}{Colors.ENDC}")
            print(border)

            log_text = (
                "=" * 40 + "\n"
                f"Flag #{flag_counter}\n"
                f"Time: {timestamp}\n"
                f"File: {file_path}\n"
                f"Type: {name}\n"
                f"Line: {line_num}\n"
                f"Offset: {offset}\n"
                f"Flag: {flag_text}\n"
                + (f"Decoded: {decoded}\n" if decoded else "")
                + f"Context:\n{ctx}\n" + "=" * 40
            )
            log_flag_to_file(log_text)
            
            # Add flag information to the list
            found_flags.append({
                'counter': flag_counter,
                'timestamp': timestamp,
                'file_path': file_path,
                'type': name,
                'line': line_num,
                'offset': offset,
                'flag': flag_text,
                'decoded': decoded,
                'context': ctx
            })

def search_lines(path, line_search_rules):
    """Search for lines that contain all specified strings"""
    try:
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            lines = f.readlines()
            for line_num, line in enumerate(lines, 1):
                for rule in line_search_rules:
                    rule_name = rule.get('name', 'Unknown Rule')
                    includes = rule.get('includes', [])
                    if all(include in line for include in includes):
                        found_lines.append({
                            'file_path': path,
                            'line_num': line_num,
                            'line': line.strip(),
                            'rule': rule_name
                        })
    except Exception:
        pass

def scan_file(path, ignore, rules, max_size, ctx_bytes, line_search_rules):
    try:
        abs_path = os.path.abspath(path)
        if abs_path in ignore or not os.path.isfile(abs_path):
            return
        size = os.path.getsize(abs_path)
        if size == 0 or size > max_size:
            return
        scanned_files[abs_path] = time.time()
        # 搜索flag
        with open(abs_path, 'rb') as f:
            try:
                with mmap.mmap(f.fileno(), 0, access=mmap.ACCESS_READ) as mm:
                    _report_flag(mm, path, ctx_bytes, rules)
            except:
                _report_flag(f.read(), path, ctx_bytes, rules)
        # 搜索符合条件的行
        search_lines(path, line_search_rules)
    except Exception:
        pass

def clean_text(text):
    """Clean text to remove characters not supported in Excel"""
    if text is None:
        return ''
    if isinstance(text, bytes):
        try:
            text = text.decode('utf-8', errors='replace')
        except:
            text = repr(text)
    # Remove any characters that are not supported in Excel
    # Excel doesn't support control characters (0x00-0x1F except 0x09, 0x0A, 0x0D)
    result = []
    for c in text:
        code = ord(c)
        # Allow printable characters and tabs, newlines, carriage returns
        if code >= 32 or code in (9, 10, 13):
            result.append(c)
    return ''.join(result)

def save_flags_to_xlsx():
    if not found_flags and not found_lines:
        log_info("没有找到flag和符合条件的行，跳过保存到xlsx")
        return
    
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # 添加Flags工作表
        ws_flags = wb.active
        ws_flags.title = "找到的Flags"
        
        # Set headers for flags
        headers_flags = ['序号', '时间戳', '文件路径', '类型', '行号', '偏移量', 'Flag', '解码', '上下文']
        for col_num, header in enumerate(headers_flags, 1):
            col_letter = get_column_letter(col_num)
            ws_flags[f'{col_letter}1'] = header
        
        # Add flag data
        for row_num, flag in enumerate(found_flags, 2):
            try:
                ws_flags[f'A{row_num}'] = flag['counter']
                ws_flags[f'B{row_num}'] = clean_text(flag['timestamp'])
                ws_flags[f'C{row_num}'] = clean_text(flag['file_path'])
                ws_flags[f'D{row_num}'] = clean_text(flag['type'])
                ws_flags[f'E{row_num}'] = flag['line']
                ws_flags[f'F{row_num}'] = flag['offset']
                ws_flags[f'G{row_num}'] = clean_text(flag['flag'])
                ws_flags[f'H{row_num}'] = clean_text(flag['decoded']) if flag['decoded'] else ''
                ws_flags[f'I{row_num}'] = clean_text(flag['context'])
            except Exception as e:
                log_warn(f"保存第{flag['counter']}个flag时出错: {e}")
                # Skip this flag and continue
                continue
        
        # Auto-adjust column widths for flags
        for column in ws_flags.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_flags.column_dimensions[column_letter].width = adjusted_width
        
        # 添加符合条件的行工作表
        if found_lines:
            ws_lines = wb.create_sheet(title="符合条件的行")
            # Set headers for lines
            headers_lines = ['文件路径', '行号', '规则', '内容']
            for col_num, header in enumerate(headers_lines, 1):
                col_letter = get_column_letter(col_num)
                ws_lines[f'{col_letter}1'] = header
            
            # Add line data
            for row_num, line_data in enumerate(found_lines, 2):
                try:
                    ws_lines[f'A{row_num}'] = clean_text(line_data['file_path'])
                    ws_lines[f'B{row_num}'] = line_data['line_num']
                    ws_lines[f'C{row_num}'] = clean_text(line_data['rule'])
                    ws_lines[f'D{row_num}'] = clean_text(line_data['line'])
                except Exception as e:
                    log_warn(f"保存第{row_num-1}行数据时出错: {e}")
                    # Skip this line and continue
                    continue
            
            # Auto-adjust column widths for lines
            for column in ws_lines.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_lines.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        xlsx_file = f"found_flags_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(xlsx_file)
        log_info(f"已将{len(found_flags)}个flag和{len(found_lines)}行符合条件的数据保存到{os.path.abspath(xlsx_file)}")
    except Exception as e:
        import traceback
        log_error(f"保存到xlsx文件失败: {e}")
        log_error(traceback.format_exc())

def initial_scan(dir_path, ignore, rules, max_size, ctx_bytes, line_search_rules):
    log_info(f"开始初始扫描: {dir_path}")
    for root, _, files in os.walk(dir_path):
        for name in files:
            scan_file(os.path.join(root, name), ignore, rules, max_size, ctx_bytes, line_search_rules)
    log_info("初始扫描完成")
    save_flags_to_xlsx()

def process_event(path, ignore, rules, max_size, ctx_bytes, cooldown, line_search_rules):
    now = time.time()
    abs_path = os.path.abspath(path)
    if abs_path in ignore or not os.path.exists(abs_path):
        return
    if os.path.isfile(abs_path):
        last = scanned_files.get(abs_path, 0)
        if now - last > cooldown:
            scan_file(abs_path, ignore, rules, max_size, ctx_bytes, line_search_rules)
    else:
        for root, _, files in os.walk(abs_path):
            for name in files:
                file_path = os.path.join(root, name)
                last = scanned_files.get(file_path, 0)
                if now - last > cooldown:
                    scan_file(file_path, ignore, rules, max_size, ctx_bytes, line_search_rules)

class WatchHandler(FileSystemEventHandler):
    def __init__(self, ignore, rules, max_size, ctx_bytes, cooldown, line_search_rules):
        super().__init__()
        self.ignore = ignore
        self.rules = rules
        self.max_size = max_size
        self.ctx_bytes = ctx_bytes
        self.cooldown = cooldown
        self.line_search_rules = line_search_rules

    def on_created(self, event):
        log_info(f"检测到创建: {event.src_path}")
        process_event(event.src_path, self.ignore, self.rules, self.max_size, self.ctx_bytes, self.cooldown, self.line_search_rules)

    def on_modified(self, event):
        if not event.is_directory:
            log_info(f"检测到修改: {event.src_path}")
            process_event(event.src_path, self.ignore, self.rules, self.max_size, self.ctx_bytes, self.cooldown, self.line_search_rules)

    def on_moved(self, event):
        log_info(f"检测到移动: {event.dest_path}")
        process_event(event.dest_path, self.ignore, self.rules, self.max_size, self.ctx_bytes, self.cooldown, self.line_search_rules)

if __name__ == "__main__":
    if len(sys.argv) >= 2:
        watch_dir = sys.argv[1]
    else:
        watch_dir = "."

    if IS_WIN:
        os.system("color")

    os.system("clear" if not IS_WIN else "cls")
    banner()

    if not os.path.isdir(watch_dir):
        log_error("提供的路径无效。")
        sys.exit(1)

    log_info("CTF Flag 搜寻器启动中...")

    config = load_or_create_config()
    rules = build_regex_from_rules(config['flag_rules'])
    line_search_rules = config.get('line_search_rules', [])
    cooldown = config['scan_cooldown']
    ctx_bytes = config['context_bytes']
    max_size = config['max_file_size_mb'] * 1024 * 1024

    # 获取脚本自身的绝对路径
    script_path = os.path.abspath(sys.argv[0])
    ignore = {
        os.path.abspath(CONFIG_FILE),
        os.path.abspath(LOG_FILE),
        script_path
    }
    log_info(f"忽略文件: {script_path}")

    log_info(f"忽略文件数: {len(ignore)}")
    initial_scan(watch_dir, ignore, rules, max_size, ctx_bytes, line_search_rules)

    log_info(f"实时监控中: {watch_dir}")
    log_info(f"按 Ctrl+C 退出。")

    handler = WatchHandler(ignore, rules, max_size, ctx_bytes, cooldown, line_search_rules)
    observer = Observer()
    observer.schedule(handler, watch_dir, recursive=True)
    observer.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        print()
        log_info("监控器已停止。")
    observer.join()
